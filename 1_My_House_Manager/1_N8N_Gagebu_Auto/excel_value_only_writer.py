import os
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook


def _normalize_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _write_dataframe_values(ws, df):
    cols = list(df.columns)
    prev_max_row = ws.max_row
    prev_max_col = ws.max_column

    # Header row: value-only update (keeps style/format)
    for col_idx, col_name in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows: value-only update (keeps style/format)
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_normalize_value(value))

    # Clear stale values from previous run while preserving formatting.
    start_clear_row = len(df) + 2
    clear_col_count = max(len(cols), prev_max_col)
    for row_idx in range(start_clear_row, prev_max_row + 1):
        for col_idx in range(1, clear_col_count + 1):
            ws.cell(row=row_idx, column=col_idx, value=None)


def write_sheets_value_only(output_path, sheets_data, template_path=None):
    """
    Write sheet data by updating cell values only.

    If output_path exists, it updates values in-place to preserve workbook/sheet styles.
    If output_path does not exist and template_path exists, template is copied first.
    If neither exists, a new workbook is created.
    """
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        if template_path and os.path.exists(template_path):
            shutil.copy2(template_path, output_path)
            wb = load_workbook(output_path)
        else:
            wb = Workbook()

    for sheet_name, df in sheets_data.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        _write_dataframe_values(ws, df)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])

    wb.save(output_path)
